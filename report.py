import os
import datetime as dt
import smtplib
from email.message import EmailMessage

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    REPORT_DIR,
    EMAIL_SKUS,
    SEND_EMAIL, SMTP_HOST, SMTP_PORT,
    SMTP_USER, SMTP_PASSWORD, EMAIL_FROM, EMAIL_TO,
)

# Colour palette
NAVY    = "1F3A5F"   # header / brand
NAVY2   = "2E5C8A"   # sub-header
RED     = "C0392B"   # critical alert
RED_LT  = "FADBD8"   # critical alert row fill
AMBER   = "D4820A"   # warning
AMB_LT  = "FDEBD0"   # warning row fill
GREEN   = "1E8449"   # positive
GRN_LT  = "D5F5E3"   # positive row fill
BLUE_LT = "EAF1F8"   # summary label fill
GREY    = "F2F3F4"   # zebra stripe
WHITE   = "FFFFFF"
BLACK   = "000000"

THIN_SIDE  = Side(style="thin",   color="BDC3C7")
MED_SIDE   = Side(style="medium", color="1F3A5F")
THIN_B     = Border(left=THIN_SIDE, right=THIN_SIDE,
                    top=THIN_SIDE,  bottom=THIN_SIDE)
MED_BOT    = Border(left=THIN_SIDE, right=THIN_SIDE,
                    top=THIN_SIDE,  bottom=MED_SIDE)

def _cell(ws, row, col, value="", font_size=9, bold=False, color=BLACK,
          bg=None, align="left", fmt=None, border=THIN_B, wrap=False,
          italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font("Calibri", size=font_size, bold=bold,
                       color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap, indent=(1 if align == "left" else 0))
    c.border    = border
    if bg:
        c.fill  = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c


def _sheet_header(ws, title, subtitle, run_date, ncols):
    """Top banner: logo area + title + date."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Row 1 – brand bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="DAILY INVENTORY REPORT")
    c.font      = Font("Calibri", bold=True, size=11, color=WHITE)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 20

    # Row 2 – title + date right-aligned
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols - 1)
    t = ws.cell(row=2, column=1, value=title)
    t.font      = Font("Calibri", bold=True, size=18, color=WHITE)
    t.fill      = PatternFill("solid", start_color=NAVY2)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    d = ws.cell(row=2, column=ncols,
                value=run_date.strftime("%b %d, %Y"))
    d.font      = Font("Calibri", bold=True, size=12, color=WHITE)
    d.fill      = PatternFill("solid", start_color=NAVY2)
    d.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[2].height = 32

    # Row 3 – subtitle
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    s = ws.cell(row=3, column=1, value=subtitle)
    s.font      = Font("Calibri", size=9, italic=True, color="5D6D7E")
    s.fill      = PatternFill("solid", start_color="D6EAF8")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[3].height = 16


def _table(wb, name, title, subtitle, frame, cols, fmts=None,
           hdr_color=NAVY, run_date=None, flag_col=None, flag_colors=None):
    """
    Write a styled data table onto a new sheet.
    flag_col   : column key whose value triggers a row colour
    flag_colors: dict mapping True/False to bg hex string
    """
    ws = wb.create_sheet(name)
    _sheet_header(ws, title, subtitle,
                  run_date or dt.date.today(), len(cols))

    HDR = 5
    ws.row_dimensions[HDR].height = 28
    for i, (_, label) in enumerate(cols, 1):
        c = ws.cell(row=HDR, column=i, value=label)
        c.font      = Font("Calibri", bold=True, color=WHITE, size=9)
        c.fill      = PatternFill("solid", start_color=hdr_color)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = MED_BOT
    ws.freeze_panes = ws.cell(row=HDR + 1, column=2)  # freeze col A + header

    row = HDR + 1
    for _, rec in frame.iterrows():
        is_disc = bool(rec.get("is_discontinued", False))
        # determine row background
        if flag_col and flag_colors:
            row_bg = flag_colors.get(bool(rec.get(flag_col)), None)
        elif is_disc:
            row_bg = "EAECEE"   # muted grey for discontinued
        else:
            row_bg = GREY if (row - HDR) % 2 == 0 else None

        ws.row_dimensions[row].height = 16
        for i, (key, _) in enumerate(cols, 1):
            val = rec.get(key, "")
            if isinstance(val, float) and pd.isna(val): val = ""
            if isinstance(val, bool): val = "Yes" if val else ""
            font_color = "95A5A6" if is_disc else BLACK
            c = ws.cell(row=row, column=i, value=val)
            c.font   = Font("Calibri", size=9, color=font_color,
                            italic=is_disc)
            c.border = THIN_B
            c.alignment = Alignment(horizontal=(
                "right" if fmts and key in fmts else "left"),
                vertical="center", indent=1)
            if row_bg:
                c.fill = PatternFill("solid", start_color=row_bg)
            if fmts and key in fmts:
                c.number_format = fmts[key]
        row += 1

    # auto column widths — money columns guaranteed min 16 to avoid #####
    MONEY_KEYS = {"inventory_value","excess_inv_value","gp_per_sku",
                  "blended_price","cost","msrp","wholesale_price"}
    for i, (key, label) in enumerate(cols, 1):
        vals  = [str(rec.get(key, "")) for _, rec in frame.iterrows()]
        width = max([len(str(label))] + [len(v) for v in vals] or [0])
        min_w = 16 if key in MONEY_KEYS else 10
        ws.column_dimensions[get_column_letter(i)].width = \
            min(max(width + 2, min_w), 40)

    if frame.empty:
        c = ws.cell(row=HDR + 1, column=1,
                    value="No items in this category today.")
        c.font = Font("Calibri", size=9, italic=True, color="95A5A6")
    return ws


def _kpi_card(ws, row, col, label, value, fmt, color, label_width=4):
    """Write one KPI card: label cols | value col."""
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + label_width - 1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font      = Font("Calibri", size=10, bold=True, color="2C3E50")
    lc.fill      = PatternFill("solid", start_color=BLUE_LT)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    lc.border    = THIN_B
    # value cell — always wide enough for currency (col + label_width)
    vc = ws.cell(row=row, column=col + label_width, value=value)
    vc.font          = Font("Calibri", size=13, bold=True, color=WHITE)
    vc.fill          = PatternFill("solid", start_color=color)
    vc.alignment     = Alignment(horizontal="center", vertical="center")
    vc.number_format = fmt
    vc.border        = THIN_B
    ws.row_dimensions[row].height = 26


def build_summary(wb, df, reorder, overstock, run_date):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    INT   = "#,##0"
    MONEY = "$#,##0"

    n_active  = int((~df["is_discontinued"]).sum())
    n_disc    = int(df["is_discontinued"].sum())
    n_reorder = len(reorder)
    n_over    = len(overstock)
    tot_units = int(df["on_hand"].sum())
    b2b_units = int(df["qty_zoey"].sum())
    d2c_units = int(df["qty_shopify"].sum())
    inv_val   = round(df["inventory_value"].sum(), 0)
    exc_val   = round(df["excess_inv_value"].sum(), 0)
    gp_val    = round(df["gp_per_sku"].sum(), 0)
    units_4w  = int(df["units_4w"].sum())
    units_13w = int(df["units_13w"].sum())
    units_52w = int(df["units_52w"].sum())
    # platform revenue
    rev_zoey_30d    = round(df["rev_zoey_30d"].sum(), 0)    if "rev_zoey_30d"    in df else 0
    rev_shopify_30d = round(df["rev_shopify_30d"].sum(), 0) if "rev_shopify_30d" in df else 0
    rev_total_30d   = round(df["rev_total_30d"].sum(), 0)   if "rev_total_30d"   in df else 0
    rev_zoey_7d     = round(df["rev_zoey_7d"].sum(), 0)     if "rev_zoey_7d"     in df else 0
    rev_shopify_7d  = round(df["rev_shopify_7d"].sum(), 0)  if "rev_shopify_7d"  in df else 0
    rev_total_7d    = round(df["rev_total_7d"].sum(), 0)    if "rev_total_7d"    in df else 0

    # ── Column widths — set FIRST so currency cells never show ##### ────
    # Layout: A-E = left label block (col 1-5), F = left value (col 6)
    #         G-K = right label block (col 7-11), L = right value (col 12)
    col_widths = {"A": 30, "B": 4, "C": 4, "D": 4, "E": 4,
                  "F": 18,
                  "G": 30, "H": 4, "I": 4, "J": 4, "K": 4,
                  "L": 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: thin accent bar ───────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 6
    ws.cell(row=1, column=1).fill = PatternFill("solid", start_color=NAVY2)

    # ── Row 2: big title + date ──────────────────────────────────────────
    ws.merge_cells("A2:I2")
    t = ws.cell(row=2, column=1, value="Daily Inventory Report")
    t.font      = Font("Calibri", bold=True, size=24, color=WHITE)
    t.fill      = PatternFill("solid", start_color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells("J2:L2")
    d = ws.cell(row=2, column=10, value=run_date.strftime("%A, %B %d, %Y"))
    d.font      = Font("Calibri", bold=True, size=11, color=WHITE)
    d.fill      = PatternFill("solid", start_color=NAVY)
    d.alignment = Alignment(horizontal="right", vertical="center", indent=2)
    ws.row_dimensions[2].height = 42

    # ── Row 3: subtitle ─────────────────────────────────────────────────
    ws.merge_cells("A3:L3")
    s = ws.cell(row=3, column=1,
                value="Zoey (B2B) + Shopify (D2C)   ·   "
                      "All figures pulled live from both platforms   ·   "
                      "Discontinued products included but excluded from alerts")
    s.font      = Font("Calibri", size=9, italic=True, color="5D6D7E")
    s.fill      = PatternFill("solid", start_color="D6EAF8")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[3].height = 18

    # spacer
    ws.row_dimensions[4].height = 10

    # ── helper: section header row ──────────────────────────────────────
    def _section(row, label):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=12)
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font("Calibri", bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 20

    # ── ALERTS ──────────────────────────────────────────────────────────
    _section(5, "⚠   ALERTS")
    _kpi_card(ws, 6, 1, "Products Needing Reorder",  n_reorder, INT,   RED,   label_width=5)
    _kpi_card(ws, 6, 7, "Products Overstocked",       n_over,    INT,   AMBER, label_width=5)
    ws.row_dimensions[7].height = 10

    # ── INVENTORY ───────────────────────────────────────────────────────
    _section(8, "📦   INVENTORY")
    _kpi_card(ws,  9, 1, "Total Products Tracked",    len(df),    INT,   NAVY2,    label_width=5)
    _kpi_card(ws,  9, 7, "Active Products",           n_active,   INT,   NAVY2,    label_width=5)
    _kpi_card(ws, 10, 1, "Total Units On Hand",       tot_units,  INT,   GREEN,    label_width=5)
    _kpi_card(ws, 10, 7, "Discontinued Products",     n_disc,     INT,   "7F8C8D", label_width=5)
    _kpi_card(ws, 11, 1, "Units on Zoey (B2B)",       b2b_units,  INT,   NAVY2,    label_width=5)
    _kpi_card(ws, 11, 7, "Units on Shopify (D2C)",    d2c_units,  INT,   NAVY2,    label_width=5)
    ws.row_dimensions[12].height = 10

    # ── FINANCIALS ──────────────────────────────────────────────────────
    _section(13, "💰   FINANCIALS  (inventory on hand)")
    _kpi_card(ws, 14, 1, "Total Inventory Value",     inv_val, MONEY, GREEN, label_width=5)
    _kpi_card(ws, 14, 7, "Total GP on Current Stock", gp_val,  MONEY, GREEN, label_width=5)
    _kpi_card(ws, 15, 1, "Excess Inventory Value",    exc_val, MONEY, AMBER, label_width=5)
    ws.row_dimensions[16].height = 10

    # ── REVENUE ─────────────────────────────────────────────────────────
    _section(17, "💵   REVENUE  (net revenue from orders)")
    _kpi_card(ws, 18, 1, "Total Revenue — Last 7 Days",      rev_total_7d,   MONEY, GREEN, label_width=5)
    _kpi_card(ws, 18, 7, "Total Revenue — Last 30 Days",     rev_total_30d,  MONEY, GREEN, label_width=5)
    _kpi_card(ws, 19, 1, "   Zoey (B2B) — Last 7 Days",     rev_zoey_7d,    MONEY, NAVY2, label_width=5)
    _kpi_card(ws, 19, 7, "   Zoey (B2B) — Last 30 Days",    rev_zoey_30d,   MONEY, NAVY2, label_width=5)
    _kpi_card(ws, 20, 1, "   Shopify (D2C) — Last 7 Days",  rev_shopify_7d, MONEY, NAVY2, label_width=5)
    _kpi_card(ws, 20, 7, "   Shopify (D2C) — Last 30 Days", rev_shopify_30d,MONEY, NAVY2, label_width=5)
    ws.row_dimensions[21].height = 10

    # ── SALES VELOCITY ──────────────────────────────────────────────────
    _section(22, "📈   SALES VELOCITY  (combined B2B + D2C)")
    _kpi_card(ws, 23, 1, "Units Sold — Last 4 Weeks",  units_4w,  INT, NAVY2, label_width=5)
    _kpi_card(ws, 23, 7, "Units Sold — Last 13 Weeks", units_13w, INT, NAVY2, label_width=5)
    _kpi_card(ws, 24, 1, "Units Sold — Last 52 Weeks", units_52w, INT, NAVY2, label_width=5)
    ws.row_dimensions[25].height = 10

    # ── Formula legend ──────────────────────────────────────────────────
    ws.merge_cells("A26:L26")
    leg = ws.cell(row=26, column=1,
        value="Formulas:  DailyVel = 52W ÷ 364   "
              "ReorderFlag = SuggestedQty>0   "
              "OverstockFlag = MonthsCover>12   "
              "MonthsCover = DaysCover÷28   "
              "Blended = MSRP×30% + Wholesale×70%   "
              "Weeks start Sunday")
    leg.font      = Font("Calibri", size=8, italic=True, color="7F8C8D")
    leg.fill      = PatternFill("solid", start_color="EBF5FB")
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[26].height = 16


def build_report(df, trend, daily, run_date, out_path):
    wb  = Workbook()
    INT, MONEY, DEC, VEL = "#,##0", "$#,##0.00", "0.0", "0.000"

    reorder   = df[~df["is_discontinued"] & df["reorder_flag"]   & (df["daily_vel"] > 0)]
    overstock = df[~df["is_discontinued"] & df["overstock_flag"] & (df["daily_vel"] > 0)]

    # shared column sets
    BASE  = [("sku","SKU"), ("name","Product")]
    INV   = [("qty_zoey","Qty Zoey"), ("qty_shopify","Qty Shopify"),
             ("on_hand","Available")]
    VEL_  = [("units_4w","Units 4W"), ("units_13w","Units 13W"),
             ("units_52w","Units 52W"),
             ("daily_vel","DailyVel"), ("days_cover","DaysCover"),
             ("months_cover","MonthsCover")]
    ORDER = [("suggested_order_qty","SuggestedOrderQty")]
    VAL   = [("blended_price","Blended Price"),
             ("inventory_value","InvValue"),
             ("excess_inv_value","ExcessInvValue"),
             ("gp_per_sku","GP Per SKU")]

    MF = {"qty_zoey":INT,"qty_shopify":INT,"on_hand":INT,
          "units_4w":INT,"units_13w":INT,"units_52w":INT,"daily_vel":VEL,
          "days_cover":DEC,"months_cover":DEC,
          "suggested_order_qty":INT,
          "blended_price":MONEY,"inventory_value":MONEY,
          "excess_inv_value":MONEY,"gp_per_sku":MONEY,
          "cost":MONEY,"msrp":MONEY,"wholesale_price":MONEY,
          "lead_time_days":INT,"target_units":INT,
          "daily_vel_4w":VEL,"daily_vel_13w":VEL}

    # ── TAB 1: Summary (enterprise dashboard) ──────────────────────────
    build_summary(wb, df, reorder, overstock, run_date)

    # ── TAB 2: Dashboard ───────────────────────────────────────────────
    _table(wb, "Dashboard", "Dashboard — All Products",
           f"{len(df)} total products sorted by urgency  ·  "
           f"{int(df['is_discontinued'].sum())} discontinued shown (greyed) but excluded from alerts",
           df,
           BASE + INV + VEL_ + ORDER
           + [("overstock_flag","OverstockFlag"),("reorder_flag","ReorderFlag"),
              ("is_discontinued","Discontinued?")],
           fmts=MF, run_date=run_date)

    # ── TAB 3: Reorder Now ─────────────────────────────────────────────
    _table(wb, "Reorder Now", "Action Required — Reorder",
           f"{len(reorder)} products flagged for reorder  ·  "
           "SuggestedOrderQty > 0  ·  Order before stock runs out",
           reorder,
           BASE + INV + VEL_ + ORDER
           + [("lead_time_days","Lead Time (days)"),
              ("target_units","Target Units")],
           fmts=MF, hdr_color=RED, run_date=run_date,
           flag_col="reorder_flag", flag_colors={True: RED_LT})

    # ── TAB 4: Overstock ───────────────────────────────────────────────
    _table(wb, "Overstock", "Overstocked Products",
           f"{len(overstock)} products with MonthsCover > 12  ·  "
           "Capital tied up — consider promotions or pausing reorders",
           overstock.sort_values("months_cover", ascending=False),
           BASE + INV
           + [("daily_vel","DailyVel"),("months_cover","MonthsCover")],
           fmts=MF, hdr_color=AMBER, run_date=run_date,
           flag_col="overstock_flag", flag_colors={True: AMB_LT})

    # ── TAB 5: Full Metrics ────────────────────────────────────────────
    _table(wb, "Full Metrics", "Full Metrics — All Products",
           f"All {len(df)} products with every metric  ·  "
           f"{int(df['is_discontinued'].sum())} discontinued included and flagged",
           df,
           BASE
           + [("qty_zoey","OnHand B2B"),("qty_shopify","OnHand D2C"),
              ("on_hand","OnHand Total"),
              ("units_4w","Units_4W"),("units_13w","Units_13W"),
              ("units_52w","Units_52W"),
              ("daily_vel_4w","DailyVel_4W"),("daily_vel_13w","DailyVel_13W"),
              ("daily_vel","DailyVel"),
              ("days_cover","DaysCover"),("months_cover","MonthsCover"),
              ("lead_time_days","LeadTimeDays"),("target_units","TargetUnits"),
              ("suggested_order_qty","SuggestedOrderQty"),
              ("reorder_flag","ReorderFlag"),("overstock_flag","OverstockFlag"),
              ("is_discontinued","Discontinued?")],
           fmts=MF, run_date=run_date)

    # ── TAB 6: Weekly Trend ────────────────────────────────────────────
    trend_df, week_labels = trend
    MONEY_K = "$#,##0"
    # merge current on-hand qty per platform into the trend frame
    trend_df = trend_df.merge(
        df[["sku","qty_zoey","qty_shopify"]], on="sku", how="left")
    trend_df["qty_zoey"]    = trend_df["qty_zoey"].fillna(0)
    trend_df["qty_shopify"] = trend_df["qty_shopify"].fillna(0)

    wt_cols = [("sku","SKU"), ("name","Product"),
               ("qty_zoey","Qty Zoey"), ("qty_shopify","Qty Shopify")]
    wt_fmts = {"qty_zoey": INT, "qty_shopify": INT}
    for wl in week_labels:
        wt_cols += [
            (f"{wl} Units",       f"{wl}\nUnits"),
            (f"{wl} Zoey Rev",    f"{wl}\nZoey Rev"),
            (f"{wl} Shopify Rev", f"{wl}\nShopify Rev"),
            (f"{wl} Net Rev",     f"{wl}\nTotal Rev"),
            (f"{wl} GP",          f"{wl}\nGP"),
        ]
        wt_fmts[f"{wl} Units"]       = INT
        wt_fmts[f"{wl} Zoey Rev"]    = MONEY_K
        wt_fmts[f"{wl} Shopify Rev"] = MONEY_K
        wt_fmts[f"{wl} Net Rev"]     = MONEY_K
        wt_fmts[f"{wl} GP"]          = MONEY_K
    for lbl, fmt in [("13W Units",INT),("13W Zoey Rev",MONEY_K),
                     ("13W Shopify Rev",MONEY_K),("13W Net Rev",MONEY_K),
                     ("13W GP",MONEY_K)]:
        wt_cols.append((lbl, lbl))
        wt_fmts[lbl] = fmt

    _table(wb, "Weekly Trend", "Weekly Sales Trend",
           f"Units · Zoey Rev · Shopify Rev · Total Net Rev · GP per week "
           f"(last {len(week_labels)} weeks, week starts Sunday) · Combined B2B + D2C",
           trend_df, wt_cols, fmts=wt_fmts, run_date=run_date)

    # ── TAB 7: Daily Sales ────────────────────────────────────────────
    _build_daily_sales_sheet(wb, daily, run_date)

    wb.save(out_path)
    print(f"Report   : {out_path}")

def _build_daily_sales_sheet(wb, daily_data, run_date):
    """
    Daily Sales — simple transaction list, last 30 days.
    Date | SKU | Product | Quantity | Subtotal | Discount Amount |
    Subtotal with Discount | Tax Amount | Cost
    One row per SKU per day that had sales. Newest first.
    """
    if daily_data is None:
        return

    df, summary, date_labels, date_strs, metric_map = daily_data
    INT   = "#,##0"
    MONEY = "$#,##0.00"

    ws = wb.create_sheet("Daily Sales")
    ws.sheet_view.showGridLines     = False
    ws.sheet_view.showRowColHeaders = False

    # ── Build flat transaction rows ──────────────────────────────────────
    rows = []
    for dl, ds in zip(date_labels, date_strs):
        for _, rec in df.iterrows():
            qty = rec.get(f"{dl}|qty", 0) or 0
            if qty <= 0:
                continue
            rows.append({
                "sort_date":       ds,  # The raw mathematical date (e.g., 2026-05-31)
                "date":            dl,
                "sku":             rec["sku"],
                "name":            rec.get("name", ""),
                "qty":             int(qty),
                "qty_zoey":        int(rec.get(f"{dl}|qty_zoey",    0) or 0),
                "qty_shopify":     int(rec.get(f"{dl}|qty_shopify", 0) or 0),
                "subtotal":        round(rec.get(f"{dl}|subtotal",  0) or 0, 2),
                "discount":        round(rec.get(f"{dl}|discount",  0) or 0, 2),
                "subtotal_w_disc": round(rec.get(f"{dl}|net_rev",   0) or 0, 2),
                "tax":             round(rec.get(f"{dl}|tax",       0) or 0, 2),
                "cost":            round(rec.get(f"{dl}|cost",      0) or 0, 2),
                "is_discontinued": bool(rec.get("is_discontinued", False)),
            })

    txn = (pd.DataFrame(rows) if rows
           else pd.DataFrame(columns=["date","sku","name","qty","qty_zoey",
                                      "qty_shopify","subtotal","discount",
                                      "subtotal_w_disc","tax","cost",
                                      "is_discontinued"]))
    txn = txn.sort_values(["sort_date","sku"], ascending=[False,True]).reset_index(drop=True)

    cols = [
        ("date",            "Date"),
        ("sku",             "SKU"),
        ("name",            "Product"),
        ("qty",             "Quantity"),
        ("qty_zoey",        "Qty Zoey"),
        ("qty_shopify",     "Qty Shopify"),
        ("subtotal",        "Subtotal"),
        ("discount",        "Discount Amount"),
        ("subtotal_w_disc", "Subtotal with Discount"),
        ("tax",             "Tax Amount"),
        ("cost",            "Cost"),
    ]
    fmts = {"qty": INT, "qty_zoey": INT, "qty_shopify": INT,
            "subtotal": MONEY, "discount": MONEY,
            "subtotal_w_disc": MONEY, "tax": MONEY, "cost": MONEY}
    widths = {"date": 12, "sku": 16, "name": 30, "qty": 11,
              "qty_zoey": 11, "qty_shopify": 12,
              "subtotal": 14, "discount": 17, "subtotal_w_disc": 22,
              "tax": 13, "cost": 13}

    ncols = len(cols)
    _sheet_header(ws, "Daily Sales",
                  f"Daily transactions since Jan 1, 2026  ·  "
                  "Combined B2B + D2C  ·  One row per product per day  ·  Newest first",
                  run_date, ncols)

    # ── Totals row ───────────────────────────────────────────────────────
    HDR = 5
    ws.row_dimensions[HDR].height = 22
    totals = {
        "date":            "ALL DAYS",
        "sku":             f"{len(txn):,} transactions",
        "name":            "",
        "qty":             int(txn["qty"].sum())             if not txn.empty else 0,
        "qty_zoey":        int(txn["qty_zoey"].sum())        if not txn.empty else 0,
        "qty_shopify":     int(txn["qty_shopify"].sum())     if not txn.empty else 0,
        "subtotal":        round(txn["subtotal"].sum(), 2)   if not txn.empty else 0,
        "discount":        round(txn["discount"].sum(), 2)   if not txn.empty else 0,
        "subtotal_w_disc": round(txn["subtotal_w_disc"].sum(),2) if not txn.empty else 0,
        "tax":             round(txn["tax"].sum(), 2)         if not txn.empty else 0,
        "cost":            round(txn["cost"].sum(), 2)        if not txn.empty else 0,
    }
    for i, (key, _) in enumerate(cols, 1):
        val = totals.get(key, "")
        c   = ws.cell(row=HDR, column=i, value=val)
        c.font      = Font("Calibri", bold=True, size=9, color=WHITE)
        c.fill      = PatternFill("solid", start_color=NAVY2)
        c.border    = MED_BOT
        c.alignment = Alignment(
            horizontal="right" if key in fmts else "left",
            vertical="center", indent=1)
        if key in fmts and val not in ("", 0):
            c.number_format = fmts[key]

    # ── Column headers ───────────────────────────────────────────────────
    HDR2 = HDR + 1
    ws.row_dimensions[HDR2].height = 22
    ws.freeze_panes = ws.cell(row=HDR2 + 1, column=1)
    for i, (key, label) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[key]
        c = ws.cell(row=HDR2, column=i, value=label)
        c.font      = Font("Calibri", bold=True, color=WHITE, size=10)
        c.fill      = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = MED_BOT

    # ── Data rows ────────────────────────────────────────────────────────
    row = HDR2 + 1
    for _, rec in txn.iterrows():
        is_disc  = bool(rec.get("is_discontinued", False))
        zebra    = "F0F3F4" if (row - HDR2) % 2 == 0 else None
        ws.row_dimensions[row].height = 16

        for i, (key, _) in enumerate(cols, 1):
            val = rec.get(key, "")
            if isinstance(val, float) and pd.isna(val): val = ""

            c = ws.cell(row=row, column=i, value=val)
            c.font      = Font("Calibri", size=9,
                               color="95A5A6" if is_disc else BLACK,
                               italic=is_disc)
            c.border    = THIN_B
            c.alignment = Alignment(
                horizontal="right" if key in fmts else "left",
                vertical="center", indent=1)
            bg = "EAECEE" if is_disc else zebra
            if bg: c.fill = PatternFill("solid", start_color=bg)
            if key in fmts and val not in ("", None, 0):
                c.number_format = fmts[key]
        row += 1

    if txn.empty:
        c = ws.cell(row=HDR2 + 1, column=1,
                    value="No sales data in the last 30 days.")
        c.font = Font("Calibri", size=9, italic=True, color="95A5A6")

# ======================================================================
# 5b.  EMAIL NOTIFICATION  -  reorder alerts for priority SKUs only
# ======================================================================

def send_reorder_email(df, run_date):
    """
    Sends a reorder-alert email to the owner.
    ONLY includes SKUs that are:
      - in EMAIL_SKUS (the owner's priority list), AND
      - not discontinued, AND
      - flagged for reorder (SuggestedOrderQty > 0, with sales velocity)
    The Excel report still contains every product — this email is the
    focused action list only.
    """
    if not SEND_EMAIL:
        print("Email   : SEND_EMAIL is false — skipping email")
        return
    if not EMAIL_TO or not SMTP_USER:
        print("Email   : EMAIL_TO / SMTP_USER not set — skipping email")
        return

    # filter to priority SKUs needing reorder
    alert = df[
        df["sku"].isin(EMAIL_SKUS)
        & ~df["is_discontinued"]
        & df["reorder_flag"]
        & (df["daily_vel"] > 0)
    ].copy()
    alert = alert.sort_values("days_cover")

    if alert.empty:
        print("Email   : no priority SKUs need reorder — no email sent")
        return

    # build HTML table of alert rows
    rows_html = ""
    for _, r in alert.iterrows():
        rows_html += (
            "<tr>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{r['sku']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{r.get('name','')}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:right'>{int(r['qty_zoey'])}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:right'>{int(r['qty_shopify'])}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:right'>{int(r['on_hand'])}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:right'>{r['days_cover']:.0f}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:right'>"
            f"<b>{int(r['suggested_order_qty'])}</b></td>"
            "</tr>"
        )

    html = f"""\
<html><body style="font-family:Calibri,Arial,sans-serif;color:#222">
  <h2 style="color:#1F3A5F;margin-bottom:4px">Reorder Alert — {run_date:%B %d, %Y}</h2>
  <p style="color:#555;margin-top:0">
    {len(alert)} priority product(s) need reordering.
    Full details for all products are in the attached report.
  </p>
  <table style="border-collapse:collapse;font-size:13px">
    <tr style="background:#1F3A5F;color:#fff">
      <th style="padding:8px 10px;text-align:left">SKU</th>
      <th style="padding:8px 10px;text-align:left">Product</th>
      <th style="padding:8px 10px">Qty Zoey</th>
      <th style="padding:8px 10px">Qty Shopify</th>
      <th style="padding:8px 10px">Available</th>
      <th style="padding:8px 10px">Days Cover</th>
      <th style="padding:8px 10px">Order Qty</th>
    </tr>
    {rows_html}
  </table>
  <p style="color:#888;font-size:11px;margin-top:14px">
    Automated reorder alert · priority SKUs only · discontinued products excluded.
  </p>
</body></html>"""

    msg = EmailMessage()
    msg["Subject"] = f"Reorder Alert — {len(alert)} products — {run_date:%b %d, %Y}"
    msg["From"]    = EMAIL_FROM
    msg["To"]      = EMAIL_TO
    msg.set_content(
        f"{len(alert)} priority products need reordering. "
        "View this email in HTML, or see the attached report.")
    msg.add_alternative(html, subtype="html")

    # attach the Excel report if it exists
    report_path = os.path.join(
        REPORT_DIR, f"Inventory_Report_{run_date:%Y-%m-%d}.xlsx")
    if os.path.exists(report_path):
        with open(report_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(report_path))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        print(f"Email   : reorder alert sent to {EMAIL_TO} "
              f"({len(alert)} priority products)")
    except Exception as e:
        print(f"WARNING  Email send failed: {e}")